from flask import Flask, render_template, redirect, jsonify, request, url_for, flash, send_from_directory, send_file
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import requests
import secrets
import string
import base64
import random
import json
import io

with open("credenziali.json", "r") as f:
    cr = json.load(f)

# generazione dell'elenco gruppi
gruppi = {"piemonte": {}, "puglia": {}}
df = pd.read_csv("gruppi_piemonte.csv")

for i in list(set(df["zona"].tolist())):
    gruppi["piemonte"][i] = []

for i in df.index:
    gruppi["piemonte"][df["zona"][i]].append(df["Denominazione Gruppo"][i])

df = pd.read_csv("gruppi_puglia.csv")

for i in list(set(df["zona"].tolist())):
    gruppi["puglia"][i] = []

for i in df.index:
    gruppi["puglia"][df["zona"][i]].append(df["Denominazione Gruppo"][i])

# costanti varie
specialita = [
    "Alpinismo",
    "Artigianato",
    "Campismo",
    "Civitas",
    "Esplorazione",
    "Espressione",
    "Giornalismo",
    "Internazionale",
    "Natura",
    "Nautica",
    "Olimpia",
    "Pronto intervento"
]

# Inizializza app e servizi
app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] =  "sqlite:///gv_db.db"
app.config["SECRET_KEY"] = secrets.token_hex()
db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = u"Sessione scaduta!"

# Classi Database
class User(db.Model, UserMixin):
    __tablename__ = "user"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(128), nullable=False, unique=True)
    password = db.Column(db.String(128), nullable=False)
    mail = db.Column(db.String(128), nullable=False)
    regione = db.Column(db.String(128), nullable=True)
    zona = db.Column(db.String(128), nullable=True)
    livello = db.Column(db.String(128), nullable=False)
    telegram_id = db.Column(db.String(128), nullable=True)

class IscrizioniEG(db.Model):
    __tablename__ = "iscrizioniEG"
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(128), nullable=False)
    stato = db.Column(db.String(128), nullable=False)
    nome = db.Column(db.String(128), nullable=False)
    mail = db.Column(db.String(128), nullable=False)
    regione = db.Column(db.String(128), nullable=False)
    zona = db.Column(db.String(128), nullable=False)
    gruppo = db.Column(db.String(128), nullable=False)
    specialita = db.Column(db.String(128), nullable=False)
    # tipo indica se conquista o conferma => True se conferma
    tipo = db.Column(db.String(128), nullable=False)
    # Contatti
    nome_capo_sq = db.Column(db.String(128), nullable=False)
    nome_capo1 = db.Column(db.String(128), nullable=False)
    mail_capo1 = db.Column(db.String(128), nullable=False)
    cell_capo1 = db.Column(db.String(128), nullable=False)
    nome_capo2 = db.Column(db.String(128), nullable=False)
    mail_capo2 = db.Column(db.String(128), nullable=False)
    cell_capo2 = db.Column(db.String(128), nullable=False)
    sesso = db.Column(db.String(128))

class WordpressUser(db.Model):
    __tablename__ = "wordpress_user"
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(128), nullable=False)
    iscrizioni_id = db.Column(db.Integer, nullable=False)
    wordpress_id = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(128), nullable=False)
    meta = db.Column(db.JSON, nullable=False)

class WordpressPost(db.Model):
    __tablename__ = "wordpress_post"
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(128), nullable=False)
    iscrizioni_id = db.Column(db.Integer, nullable=False)
    wordpress_user_id = db.Column(db.Integer, nullable=False)
    wordpress_id = db.Column(db.Integer, nullable=False)
    # campo di testo per indicare se "presentazione", "prima_impresa", "seconda_impresa", "missione", "extra"
    tipo = db.Column(db.String(128), nullable=False)
    meta = db.Column(db.JSON, nullable=False)

class RelazioniPuglia(db.Model):
    __tablename__ = "relazioni_puglia"
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(128), nullable=False)
    stato = db.Column(db.JSON, nullable=False)
    iscrizioni_id = db.Column(db.Integer, nullable=False)
    dati = db.Column(db.JSON, nullable=False)

class TestiMail(db.Model):
    __tablename__ = "testi_mail"
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(128), nullable=False)
    stato = db.Column(db.JSON, nullable=False)
    destinatari = db.Column(db.JSON, nullable=False)
    titolo = db.Column(db.String(128), nullable=False)
    testo = db.Column(db.UnicodeText, nullable=False)

class StatusPercorso(db.Model):
    __tablename__ = "status_percorso"
    id = db.Column(db.Integer, primary_key=True)
    iscrizioni = db.Column(db.JSON, nullable=False)
    abilitazioni = db.Column(db.JSON, nullable=False)
    regione = db.Column(db.String(128), nullable=True)
    data_apertura = db.Column(db.String(128), nullable=False)
    data_chiusura = db.Column(db.String(128), nullable=False)

with app.app_context():
    db.create_all()

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def manda_mail(indirizzi, copia, titolo, testo, regione="piemonte"):
    message = MIMEMultipart("alternative")
    message["Subject"] = f"Guidoncini Verdi 2025 - {titolo}"
    message["From"] = cr["mail"]["sender_email"]
    if regione == "puglia":
        message["From"] = cr["mail_puglia"]["sender_email"]
    message["To"] = ", ".join(indirizzi)
    if copia:
        message["Cc"] = ", ".join(copia)
        indirizzi.extend(copia)

    text = f"{titolo}\n{testo}"
    html = render_template("mail_base.html", titolo=titolo, testo=testo)

    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")

    message.attach(part1)
    message.attach(part2)

    context = ssl.create_default_context()
    try:
        if regione == "piemonte":
            with smtplib.SMTP(cr["mail"]["smtp_server"], cr["mail"]["port"]) as server:
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(cr["mail"]["sender_email"], cr["mail"]["passwd"])
                server.sendmail(cr["mail"]["sender_email"], indirizzi, message.as_string())
                server.quit()
        elif regione == "puglia":
            with smtplib.SMTP(cr["mail_puglia"]["smtp_server"], cr["mail_puglia"]["port"]) as server:
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(cr["mail_puglia"]["sender_email"], cr["mail_puglia"]["passwd"])
                server.sendmail(cr["mail_puglia"]["sender_email"], indirizzi, message.as_string())
                server.quit()
        return True
    except:
        return False

def manda_telegram(chat_id, titolo, testo):
    text = f"{titolo}\n{testo}"
    t_url = f"https://api.telegram.org/bot{cr['telegram']['token']}/sendMessage?chat_id={chat_id}&text={text}"
    try:
        requests.get(t_url)
        return True
    except:
        return False

def genera_password_sq():
    nomi = ["Akela", "Baloo", "Chil", "Kaa", "Raksha", "Arcanda", "Sciba", "Scoiattoli", "Mi", "Mowgli"]
    colori = ["Rosso", "Blu", "Verde", "Giallo", "Arancione", "Viola", "Rosa", "Marrone", "Grigio", "Nero"]
    return f"{random.choice(nomi)}{random.choice(colori)}"

def crea_utente(id_iscrizione, header, dati):
    try:
        response = requests.post(cr["wordpress"]["url"]+"/users", headers=header, json=dati)
        id_autore = response.json()["id"]
    except Exception as e:
        print(e)
        return False
    utente = WordpressUser(data=str(datetime.now()), iscrizioni_id=int(id_iscrizione), wordpress_id=int(id_autore), username=dati["username"], meta=dati)
    db.session.add(utente)
    db.session.commit()
    return id_autore

def crea_post(id_iscrizione, wp_id, header, dati, tipo):
    try:
        response = requests.post(cr["wordpress"]["url"]+"/posts", headers=header, json=dati)
        id_post = response.json()["id"]
    except:
        return False
    post = WordpressPost(data=str(datetime.now()), iscrizioni_id=int(id_iscrizione), wordpress_user_id=wp_id, wordpress_id=int(id_post), tipo=tipo, meta=dati)
    db.session.add(post)
    db.session.commit()
    return True

def crea_navigazione(id_iscrizione, wp_id, header, dati, tipo):
    try:
        response = requests.post(cr["wordpress"]["url"]+"/navigazione", headers=header, json=dati)
        id_post = response.json()["id"]
    except:
        return False
    post = WordpressPost(data=str(datetime.now()), iscrizioni_id=int(id_iscrizione), wordpress_user_id=wp_id, wordpress_id=int(id_post), tipo=tipo, meta=dati)
    db.session.add(post)
    db.session.commit()
    return True

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/dashboard")
@login_required
def dashboard():
    non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").count()
    if current_user.livello == "iabz":
        non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").filter_by(zona=current_user.zona).count()
        stato = StatusPercorso.query.filter_by(regione=current_user.regione).first().iscrizioni
    if current_user.livello == "iabr":
        non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").filter_by(regione=current_user.regione).count()
        stato = StatusPercorso.query.filter_by(regione=current_user.regione).first().iscrizioni
    if current_user.livello == "admin":
        stato = False
    return render_template("dashboard.html", stato=stato, non_abilitate=non_abilitate)

@app.route("/stato_iscrizioni", methods=["GET", "POST"])
@login_required
def stato_iscrizioni():
    if current_user.livello == "iabz" or current_user.livello == "pattuglia":
        return redirect(url_for("dashboard"))
    if current_user.livello == "iabr":
        stato = StatusPercorso.query.filter_by(regione=current_user.regione).first()
    if request.method == "POST":
        if request.form["stato"] == "sospendi":
            stato.iscrizioni = False
            db.session.commit()
        if request.form["stato"] == "apri":
            stato.iscrizioni = True
            stato.data_apertura = str(datetime.now())
            stato.data_chiusura = ""
            db.session.commit()
        if request.form["stato"] == "chiudi":
            stato.iscrizioni = False
            stato.data_chiusura = str(datetime.now())
            db.session.commit()
        if request.form["stato"] == "abilita":
            stato.abilitazioni = True
            db.session.commit()
        if request.form["stato"] == "ferma":
            stato.abilitazioni = False
            db.session.commit()
        return redirect(url_for("stato_iscrizioni"))
    return render_template("stato_iscrizioni.html", stato=stato)

@app.route("/iscrizioni")
@login_required
def iscrizioni():
    limita = False
    if current_user.livello == "iabz":
        limita = True
    iscritti = []
    tmp_iscritti=IscrizioniEG.query.filter_by(regione=current_user.regione)
    for i in tmp_iscritti:
        if limita and i.zona != current_user.zona:
            continue
        iscritti.append(i)
    if current_user.livello == "admin":
        iscritti = []
        tmp_iscritti=IscrizioniEG.query.all()
        for i in tmp_iscritti:
            iscritti.append(i)
    return render_template("iscrizioni.html", iscritti=iscritti)

@app.route("/report")
@login_required
def report():
    limita = False
    if current_user.livello == "iabz":
        limita = True
    iscritti = []
    tmp_iscritti=IscrizioniEG.query.filter_by(regione=current_user.regione)
    for i in tmp_iscritti:
        if limita and i.zona != current_user.zona:
            continue
        iscritti.append(i)
    wb = Workbook()
    ws = wb.active
    ws.title = "iscrizioni"
    #titoli delle colonne
    ws.cell(row=1, column=1).value = "Informazioni Cronologiche"
    ws.cell(row=1, column=2).value = "Nome Sq"
    ws.cell(row=1, column=3).value = "Gruppo"
    ws.cell(row=1, column=4).value = "Zona"
    ws.cell(row=1, column=5).value = "Specialità"
    ws.cell(row=1, column=6).value = "Tipo"
    ws.cell(row=1, column=7).value = "Nome Capo Sq"
    ws.cell(row=1, column=8).value = "Mail Capo Sq"
    ws.cell(row=1, column=9).value = "Nome Capo Rep 1"
    ws.cell(row=1, column=10).value = "Mail Capo Rep 1"
    ws.cell(row=1, column=11).value = "Cell Capo Rep 1"
    ws.cell(row=1, column=12).value = "Nome Capo Rep 2"
    ws.cell(row=1, column=13).value = "Mail Capo Rep 2"
    ws.cell(row=1, column=14).value = "Cell Capo Rep 2"
    ws.cell(row=1, column=15).value = "Stato Iscrizione"

    for i, iscritto in enumerate(iscritti):
        tmp_riga = i+2
        ws.cell(row=tmp_riga, column=1).value = iscritto.data
        ws.cell(row=tmp_riga, column=2).value = iscritto.nome
        ws.cell(row=tmp_riga, column=3).value = iscritto.gruppo
        ws.cell(row=tmp_riga, column=4).value = iscritto.zona
        ws.cell(row=tmp_riga, column=5).value = iscritto.specialita
        ws.cell(row=tmp_riga, column=6).value = iscritto.tipo
        ws.cell(row=tmp_riga, column=7).value = iscritto.nome_capo_sq
        ws.cell(row=tmp_riga, column=8).value = iscritto.mail
        ws.cell(row=tmp_riga, column=9).value = iscritto.nome_capo1
        ws.cell(row=tmp_riga, column=10).value = iscritto.mail_capo1
        ws.cell(row=tmp_riga, column=11).value = iscritto.cell_capo1
        ws.cell(row=tmp_riga, column=12).value = iscritto.nome_capo2
        ws.cell(row=tmp_riga, column=13).value = iscritto.mail_capo2
        ws.cell(row=tmp_riga, column=14).value = iscritto.cell_capo2
        ws.cell(row=tmp_riga, column=15).value = iscritto.stato

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="riepilogo.xlsx")

@app.route("/dettagli/<id_iscrizione>")
@login_required
def dettagli(id_iscrizione):
    creds = f"{cr['wordpress']['user']}:{cr['wordpress']['passwd']}"
    token = base64.b64encode(creds.encode())
    header = {"Authorization": f"Basic {token.decode('utf-8')}"}
    tmp_iscrizione = IscrizioniEG.query.filter_by(id=int(id_iscrizione)).first()
    link_sq = ""
    if tmp_iscrizione.stato == "abilitato":
        try:
            tmp_wordpress_id = WordpressPost.query.filter_by(iscrizioni_id=int(id_iscrizione)).filter_by(tipo="posts").first().wordpress_id
            link_sq = requests.get(cr["wordpress"]["url"]+"/posts/"+str(tmp_wordpress_id), headers=header).json()["link"]
        except:
            link_sq = ""
    try:
        relazione = RelazioniPuglia.query.filter_by(iscrizioni_id=int(id_iscrizione)).first()
    except:
        relazione = False
    return render_template("dettaglio_iscrizione.html", iscrizione=tmp_iscrizione, link_sq=link_sq, relazione=relazione)

@app.route("/elimina/<id_iscrizione>")
@login_required
def elimina(id_iscrizione):
    iscrizione=IscrizioniEG.query.filter_by(id=int(id_iscrizione)).first()
    try:
        if iscrizione.stato == "abilitato":
            flash("L'utente è già stato abilitato!", "warning")
            return redirect(url_for("iscrizioni"))
    except:
        flash("Non ho trovato l'iscrizione!", "warning")
        return redirect(url_for("iscrizioni"))
    iscrizione.stato = "eliminato"
    db.session.commit()
    return redirect(url_for("iscrizioni"))

@app.route("/elimina_def/<id_iscrizione>", methods=["GET", "POST"])
@login_required
def elimina_def(id_iscrizione):
    if current_user.livello != "admin":
        return redirect(url_for("dashboard"))
    iscrizione=IscrizioniEG.query.filter_by(id=int(id_iscrizione)).first()
    try:
        if iscrizione.stato == "abilitato":
            flash("L'utente è già stato abilitato!", "warning")
            return redirect(url_for("iscrizioni"))
    except:
        flash("Non ho trovato l'iscrizione!", "warning")
        return redirect(url_for("iscrizioni"))
    if request.method == "POST":
        db.session.delete(iscrizione)
        db.session.commit()
        return redirect(url_for("iscrizioni"))
    return render_template("elimina.html", iscrizione=iscrizione)

@app.route("/ripristina/<id_iscrizione>")
@login_required
def ripristina(id_iscrizione):
    iscrizione=IscrizioniEG.query.filter_by(id=int(id_iscrizione)).first()
    try:
        if iscrizione.stato == "abilitato":
            flash("L'utente è già stato abilitato!", "warning")
            return redirect(url_for("iscrizioni"))
    except:
        flash("Non ho trovato l'iscrizione!", "warning")
        return redirect(url_for("iscrizioni"))
    iscrizione.stato = "da_abilitare"
    db.session.commit()
    return redirect(url_for("iscrizioni"))

@app.route("/edit/<id_iscrizione>", methods=["GET", "POST"])
@login_required
def edit_iscrizione(id_iscrizione):
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    iscrizione=IscrizioniEG.query.filter_by(id=int(id_iscrizione)).first()
    try:
        if iscrizione.stato == "abilitato":
            flash("L'utente è già stato abilitato!", "warning")
            return redirect(url_for("iscrizioni"))
        elif iscrizione.stato == "eliminato":
            flash("Utente eliminato. Ripristinalo per poterlo modificare.", "warning")
            return redirect(url_for("iscrizioni"))
    except:
        flash("Non ho trovato l'iscrizione!", "warning")
        return redirect(url_for("iscrizioni"))
    if request.method == "POST":
        try:
            iscrizione.nome = request.form["nome_squadriglia"].capitalize()
            iscrizione.mail = request.form["mail_squadriglia"]
            iscrizione.zona = request.form["zona"]
            iscrizione.sesso = request.form["tipo_sq"]
            iscrizione.gruppo = request.form["gruppo"]
            iscrizione.specialita = request.form["specialita"]
            iscrizione.tipo = request.form["conquista_conferma"]
            iscrizione.nome_capo_sq = request.form["nome_capo_squadriglia"]
            iscrizione.nome_capo1 = request.form["nome_capo_rep1"]
            iscrizione.mail_capo1 = request.form["mail_rep1"]
            iscrizione.cell_capo1 = request.form["numero_rep1"]
            iscrizione.nome_capo2 = request.form["nome_capo_rep2"]
            iscrizione.mail_capo2 = request.form["mail_rep2"]
            iscrizione.cell_capo2 = request.form["numero_rep2"]
            db.session.commit()
        except:
            flash("Modifica Iscrizione fallita. Riprovaci!", "warning")
            return redirect(url_for("iscrizioni"))

        testo_mail_sq = f"Carə {iscrizione.nome},<br>la vostra iscrizione al percorso Guidoncini Verdi 2025 è stata modificata come richiesto.<hr><h4><strong>Dettagli Iscrizione</strong></h4>Zona: {iscrizione.zona}<br>Gruppo: {iscrizione.gruppo}<br>Ambito scelto: {iscrizione.specialita} - {iscrizione.tipo.capitalize()}"
        manda_mail([iscrizione.mail], [iscrizione.mail_capo1, iscrizione.mail_capo2], "Modifica iscrizione", testo_mail_sq)

        # Avvisa Francesco e Admin
        try:
            testo_telegram = f"Squadriglia {iscrizione.nome}\n{iscrizione.gruppo} - {iscrizione.zona}\nAmbito\n{iscrizione.specialita} - {iscrizione.tipo.capitalize()}"
            manda_telegram(User.query.filter_by(username="egm").first().telegram_id, "Modifica Iscrizione", testo_telegram)
            manda_telegram(User.query.filter_by(username="admin").first().telegram_id, "Modifica Iscrizione", testo_telegram)
        except:
            print("Errore Telegram")
        return redirect(url_for("iscrizioni"))
    return render_template("edit_iscrizione.html", iscrizione=iscrizione, gruppi=gruppi[current_user.regione], specialita=specialita)

@app.route("/abilita/<id_iscrizione>", methods=["GET", "POST"])
@login_required
def abilita(id_iscrizione):
    if not StatusPercorso.query.filter_by(regione=current_user.regione).first().abilitazioni:
        return redirect(url_for("iscrizioni"))
    creds = f"{cr['wordpress']['user']}:{cr['wordpress']['passwd']}"
    token = base64.b64encode(creds.encode())
    header = {"Authorization": f"Basic {token.decode('utf-8')}"}
    tmp_iscrizione = IscrizioniEG.query.filter_by(id=id_iscrizione).first()
    try:
        tmp_username = f"{tmp_iscrizione.nome.strip(' ')}_{tmp_iscrizione.gruppo}".replace(" ", "_").lower()
    except:
        flash("Non ho trovato l'iscrizione!", "warning")
        return redirect(url_for("iscrizioni"))
    tmp_passwd = genera_password_sq()
    response = requests.get(cr["wordpress"]["url"]+"/users", headers=header)
    valid_username = True
    for i in response.json():
        if i["slug"] == tmp_username:
            valid_username = False
    if request.method == "POST":
        try:
            tmp_username = request.form["username"]
        except KeyError:
            tmp_username = f"{tmp_iscrizione.nome.strip(' ')}_{tmp_iscrizione.gruppo}".replace(" ", "_").lower()
        valid_username = True
        for i in response.json():
            if i["slug"] == tmp_username:
                valid_username = False
        if not valid_username:
            return render_template("abilita.html", iscrizione=tmp_iscrizione, username=tmp_username, valid_username=valid_username)
        if tmp_iscrizione.tipo == "conquista":
            tmp_rinnovo = False
        else:
            tmp_rinnovo = True
        if tmp_iscrizione.zona == "ZONA DEI VINI":
            tmp_zona = "dei Vini"
        elif tmp_iscrizione.zona == "REGIONE VALLE D’AOSTA":
            tmp_zona = "Valle d'Aosta"
        else:
            tmp_zona = tmp_iscrizione.zona.removeprefix("ZONA ").capitalize()
        if tmp_iscrizione.specialita.capitalize() == "Pronto intervento":
            tmp_specialita = "Pronto Intervento"
        else:
            tmp_specialita = tmp_iscrizione.specialita.capitalize()

        tmp_meta = {
            "anno": "2025",
            "gruppo": tmp_iscrizione.gruppo.capitalize(),
            "rinnovo": tmp_rinnovo,
            "specialita": tmp_specialita,
            "squadriglia": tmp_iscrizione.nome.capitalize(),
            "regione": tmp_iscrizione.regione.capitalize(),
            "zona": tmp_zona
            }

        dati = {
            "username": tmp_username,
            "name": tmp_iscrizione.nome.capitalize(),
            "email": tmp_iscrizione.mail,
            "password": tmp_passwd,
            "meta": tmp_meta
            }

        id_autore = crea_utente(id_iscrizione, header, dati)
        if not id_autore:
            flash("Qualcosa è andato storto con la creazione dell'utente", "warning")
            return redirect(url_for("iscrizioni"))

        tmp_iscrizione.stato = "abilitato"
        db.session.commit()

        tmp_ok = True

        tmp_content = requests.get(cr["wordpress"]["url"]+"/posts/8183?context=edit", headers=header).json()["content"]["raw"]

        dati = {
            "author": int(id_autore),
            "categories": [22],
            "content": tmp_content,
            "meta": tmp_meta,
            "specialita": [specialita.index(tmp_iscrizione.specialita.capitalize())+3],
            "title": f"{tmp_meta['squadriglia']}",
            "status": "publish"
            }
        if not crea_post(id_iscrizione, id_autore, header, dati, "posts"):
            tmp_ok = False

        if not tmp_ok:
            flash("Qualcosa è andato storto con la creazione dei post, abbiamo avvisato gli IABR", "warning")
            try:
                testo_telegram = f"Squadriglia {tmp_iscrizione.nome}\n{tmp_iscrizione.gruppo} - {tmp_iscrizione.zona}\nNon tutti i post sono stati correttamente creati"
                manda_telegram(User.query.filter_by(username="egm").first().telegram_id, "Problema tecnico!!", testo_telegram)
                manda_telegram(User.query.filter_by(username="admin").first().telegram_id, "Problema tecnico!!", testo_telegram)
            except:
                print("Errore")

        testo_mail_sq = f"Congratulazioni {tmp_iscrizione.nome},<br>ecco le credenziali per il Diario di Bordo Digitale, potete accedere <a href=\"https://guidonciniverdi.it/wp-login.php\" target=\"_blank\">cliccando qui</a> oppure scaricando la app.<br><a href=\"https://play.google.com/store/apps/details?id=org.wordpress.android\" target=\"_blank\">Clicca qui per scaricare la app per Android</a><br><a href=\"https://apps.apple.com/it/app/wordpress-website-builder/id335703880\" target=\"_blank\">Clicca qui per scaricare la app per iOS</a><br>Trovate maggiori info qui: <a href=\"https://guidonciniverdi.it/come-funziona/\" target=\"_blank\">guidonciniverdi.it/come-funziona/</a><hr><h4><strong>Credenziali</strong></h4>Username: {tmp_username}<br>Password: {tmp_passwd}"
        manda_mail([tmp_iscrizione.mail], [tmp_iscrizione.mail_capo1, tmp_iscrizione.mail_capo2], "Credenziali Diario di Bordo!", testo_mail_sq, tmp_iscrizione.regione)

        return redirect(url_for("iscrizioni"))
    return render_template("abilita.html", iscrizione=tmp_iscrizione, username=tmp_username, valid_username=valid_username)

@app.route("/mail")
@login_required
def mail():
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    return render_template("testi_mail.html", testi_mail=TestiMail.query.all())

@app.route("/test_mail/<id_mail>")
@login_required
def test_mail(id_mail):
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    testo_mail = TestiMail.query.filter_by(id=id_mail).first()
    manda_mail([current_user.mail], [], testo_mail.titolo, testo_mail.testo)
    return redirect(url_for("mail"))

@app.route("/send_mail/<id_mail>")
@login_required
def send_mail(id_mail):
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    testo_mail = TestiMail.query.filter_by(id=id_mail).first()
    tmp_iscritti=IscrizioniEG.query.all()
    tmp_destinatari = []
    tmp_copia = []
    for i in tmp_iscritti:
        if i.stato == "eliminato":
            continue
        if testo_mail.destinatari["sq"] and i.stato == "da_abilitare":
            tmp_destinatari.append(i.mail)
            tmp_copia.append(i.mail_capo1)
            tmp_copia.append(i.mail_capo2)
        if testo_mail.destinatari["sq_abilitate"] and i.stato == "abilitato":
            tmp_destinatari.append(i.mail)
            tmp_copia.append(i.mail_capo1)
            tmp_copia.append(i.mail_capo2)
        if testo_mail.destinatari["capi"] and i.stato == "abilitato":
            tmp_destinatari.append(i.mail_capo1)
            tmp_destinatari.append(i.mail_capo2)
    destinatari = list(set(tmp_destinatari))
    destinatari_copia = list(set(tmp_copia))
    manda_mail(destinatari, destinatari_copia, testo_mail.titolo, testo_mail.testo)
    testo_mail.stato = True
    db.session.commit()
    return redirect(url_for("mail"))

@app.route("/delete_mail/<id_mail>")
@login_required
def delete_mail(id_mail):
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    testo_mail = TestiMail.query.filter_by(id=id_mail).first()
    db.session.delete(testo_mail)
    db.session.commit()
    return redirect(url_for("mail"))

@app.route("/crea_mail", methods=["GET", "POST"])
@login_required
def crea_mail():
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    if request.method == 'POST':
        destinatari = {"sq": False,"sq_abilitate": False, "capi": False}
        try:
            request.form["squadriglie"]
            destinatari["sq"] = True
        except KeyError:
            destinatari["sq"] = False
        try:
            request.form["squadriglie_abilitate"]
            destinatari["sq_abilitate"] = True
        except KeyError:
            destinatari["sq_abilitate"] = False
        try:
            request.form["capi_reparto"]
            destinatari["capi"] = True
        except KeyError:
            destinatari["capi"] = False
        testo_mail = TestiMail(data=str(datetime.now()), stato=False, destinatari=destinatari, titolo=request.form["titolo"], testo=request.form["ckeditor"])
        db.session.add(testo_mail)
        db.session.commit()
        return redirect(url_for("mail"))
    return render_template("testo_mail.html")

@app.route("/edit_mail/<id_mail>", methods=["GET", "POST"])
@login_required
def edit_mail(id_mail):
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    testo_mail = TestiMail.query.filter_by(id=id_mail).first()
    if request.method == 'POST':
        destinatari = {"sq": False,"sq_abilitate": False, "capi": False}
        try:
            request.form["squadriglie"]
            destinatari["sq"] = True
        except KeyError:
            destinatari["sq"] = False
        try:
            request.form["squadriglie_abilitate"]
            destinatari["sq_abilitate"] = True
        except KeyError:
            destinatari["sq_abilitate"] = False
        try:
            request.form["capi_reparto"]
            destinatari["capi"] = True
        except KeyError:
            destinatari["capi"] = False
        testo_mail.data = str(datetime.now())
        testo_mail.destinatari = destinatari
        testo_mail.stato = False
        testo_mail.titolo = request.form["titolo"]
        testo_mail.testo = request.form["ckeditor"]
        db.session.commit()
        return redirect(url_for("mail"))
    return render_template("edit_testo_mail.html", testo_mail=testo_mail)

@app.route("/login", methods=["GET", "POST"])
def login():
    if len(User.query.all()) == 0:
        return render_template("welcome.html")
    if request.method == "POST":
        utente = User.query.filter_by(username=request.form["username"]).first()
        if utente:
            if check_password_hash(utente.password, request.form["passwd"]):
                login_user(utente)
                return redirect(url_for("dashboard"))
            else:
                flash("Username o Password errati!", "warning")
        else:
            flash("Utente inesistente!", "warning")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("dashboard"))

@app.route("/admin", methods=["GET", "POST"])
@login_required
def admin():
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        if request.form["id_form"] == "nuovo_utente":
            utente = User.query.filter_by(username=request.form["username"]).first()
            if utente is None:
                alphabet = string.ascii_letters + string.digits
                tmp_password = ''.join(secrets.choice(alphabet) for i in range(12))
                password = generate_password_hash(tmp_password)
                if current_user.livello == "admin":
                    regione = request.form["regione"]
                else:
                    regione = current_user.regione
                if request.form["livello"] == "iabz":
                    utente = User(username=request.form["username"], password=password, mail=request.form["mail"], livello=request.form["livello"], regione=regione, zona=request.form["zona"], telegram_id=request.form["telegram_id"])
                else:
                    utente = User(username=request.form["username"], password=password, mail=request.form["mail"], livello=request.form["livello"], regione=regione, telegram_id=request.form["telegram_id"])
                db.session.add(utente)
                db.session.commit()
                flash("Utente inserito con successo!", "success")

                testo_mail = f"Benvenuto {utente.username},<br>la presente per confermarti la creazione dell'account sul Gestionale Guidoncini Verdi 2025 Regione {utente.regione.capitalize()}!<br>Il Gestionale è la piattaforma usata per gestire le iscrizioni dei ragazzi e il nuovissimo sito <a href=\"guidonciniverdi.it\" target=\"_blank\">guidonciniverdi.it</a>.<hr><h4><strong>Dettagli Iscrizione</strong></h4>Username: {utente.username}<br>Password provvisoria: {tmp_password}<br>Per accedere al gestionale puoi cliccare a questo <a href=\"guidonciniverdi.pythonanywhere.com/dashboard\" target=\"_blank\">link</a>"

                if manda_mail([utente.mail], [], "Conferma Creazione Account", testo_mail, utente.regione):
                    flash("Mail inviata!", "success")
                else:
                    flash("Qualcosa è andato storto con la mail...", "warning")
                if utente.telegram_id:
                    testo_telegram = f"Benvenuto {utente.username},\nla presente per confermarti la creazione dell'account sul Gestionale Guidoncini Verdi 2025!\nIl Gestionale è la piattaforma usata per gestire le iscrizioni dei ragazzi e il nuovissimo sito guidonciniverdi.it.\n\nDettagli Iscrizione\nUsername: {utente.username}\nPassword provvisoria: {tmp_password}\nPer accedere al gestionale puoi cliccare a questo link: guidonciniverdi.pythonanywhere.com/dashboard"
                    if manda_telegram(utente.telegram_id, "Conferma Creazione Account", testo_telegram):
                        flash("Notifica telegram inviata!", "success")
                    else:
                        flash("Qualcosa è andato storto con la notifica telegram...", "warning")
            else:
                flash(f"Esiste già l'utente {request.form['username']}!", "warning")
        return redirect(url_for("admin"))
    if current_user.livello == "admin":
        utenti=User.query.all()
        tmp_gruppi = gruppi["piemonte"]
    else:
        utenti=User.query.filter_by(regione=current_user.regione)
        tmp_gruppi = gruppi[current_user.regione]
    return render_template("admin.html", utenti=utenti, gruppi=tmp_gruppi)

@app.route("/crea_account")
@login_required
def crea_account():
    if (current_user.livello != "iabr") and (current_user.livello != "admin"):
        return redirect(url_for("dashboard"))
    if current_user.livello == "admin":
        utenti=User.query.all()
        tmp_gruppi = gruppi["piemonte"]
    else:
        utenti=User.query.filter_by(regione=current_user.regione)
        tmp_gruppi = gruppi[current_user.regione]
    return render_template("crea_account.html", utenti=utenti, gruppi=tmp_gruppi)

@app.route("/utente", methods=["GET", "POST"])
@login_required
def utente():
    if request.method == "POST":
        if request.form["id_form"] == "aggiorna_utente":
            utente = User.query.filter_by(username=current_user.username).first()
            if request.form["passwd"] == request.form["conferma_passwd"] and check_password_hash(utente.password, request.form["old_passwd"]):
                utente.password = generate_password_hash(request.form["passwd"])
                db.session.commit()
            else:
                flash("Le password non coincidono!", "warning")
        return redirect(url_for("utente"))
    return render_template("utente.html")

@app.route("/welcome", methods=["POST"])
def welcome():
    if len(User.query.all()) > 0:
        return redirect(url_for("login"))
    if request.form["passwd"] == request.form["conferma_passwd"]:
        password = generate_password_hash(request.form["passwd"])
        utente = User(username=request.form["username"], password=password, mail=request.form["mail"], livello="admin", telegram_id=request.form["telegram_id"])
        db.session.add(utente)
        status = StatusPercorso(iscrizioni=False, abilitazioni=False, regione="piemonte", data_apertura="", data_chiusura="")
        db.session.add(status)
        status = StatusPercorso(iscrizioni=False, abilitazioni=False, regione="puglia", data_apertura="", data_chiusura="")
        db.session.add(status)
        db.session.commit()
        flash("Utente creato con successo!", "success")
    else:
        flash("Le password non coincidono!", "warning")
    return redirect(url_for("login"))

@app.route("/iscrivi")
@login_required
def iscrivi():
    return render_template("iscrivi.html", gruppi=gruppi[current_user.regione], specialita=specialita)

@app.route("/iscriviti/<regione>", methods=["GET", "POST"])
def iscriviti(regione):
    if request.method == "POST":
        if request.form["nome_squadriglia"].replace(" ", "") == "" or request.form["mail_squadriglia"].replace(" ", "") == "" or request.form["nome_capo_squadriglia"].replace(" ", "") == "" or request.form["nome_capo_rep1"].replace(" ", "") == "" or request.form["mail_rep1"].replace(" ", "") == "" or request.form["numero_rep1"].replace(" ", "") == "":
            flash("Non hai compilato i campi obbligatori. Riprovaci!", "warning")
            return redirect(url_for("iscriviti"))
        try:
            iscrizione = IscrizioniEG(data=str(datetime.now()), stato="da_abilitare", nome=request.form["nome_squadriglia"].capitalize(), sesso=request.form["tipo_sq"], mail=request.form["mail_squadriglia"], regione=regione, zona=request.form["zona"], gruppo=request.form["gruppo"], specialita=request.form["specialita"], tipo=request.form["conquista_conferma"], nome_capo_sq=request.form["nome_capo_squadriglia"], nome_capo1=request.form["nome_capo_rep1"], mail_capo1=request.form["mail_rep1"], cell_capo1=request.form["numero_rep1"], nome_capo2=request.form["nome_capo_rep2"], mail_capo2=request.form["mail_rep2"], cell_capo2=request.form["numero_rep2"])
            db.session.add(iscrizione)
            db.session.commit()
        except:
            flash("Iscrizione fallita. Riprovaci!", "warning")
            return redirect(url_for("iscriviti"))

        testo_mail_sq = f"Congratulazioni {iscrizione.nome},<br>la vostra iscrizione al percorso Guidoncini Verdi 2025 è stata registrata!<br>Nelle prossime settimane riceverete una mail con le credenziali per accedere al vostro Diario di Bordo Digitale, nell'attesa potete iniziare a scoprire il nostro nuovissimo sito <a href=\"https://guidonciniverdi.it/\" target=\"_blank\">guidonciniverdi.it</a>.<hr><h4><strong>Dettagli Iscrizione</strong></h4>Zona: {iscrizione.zona}<br>Gruppo: {iscrizione.gruppo}<br>Ambito scelto: {iscrizione.specialita} - {iscrizione.tipo.capitalize()}"
        manda_mail([iscrizione.mail], [iscrizione.mail_capo1, iscrizione.mail_capo2], "Iscrizione completata!", testo_mail_sq, iscrizione.regione)

        # Avvisa Francesco e Admin
        try:
            testo_telegram = f"Squadriglia {iscrizione.nome}\n{iscrizione.gruppo} - {iscrizione.zona}\nAmbito\n{iscrizione.specialita} - {iscrizione.tipo.capitalize()}"
            manda_telegram(User.query.filter_by(username="iabr_piemonte").first().telegram_id, "Nuova Iscrizione", testo_telegram)
            manda_telegram(User.query.filter_by(username="admin").first().telegram_id, "Nuova Iscrizione", testo_telegram)
        except:
            print("Errore")
        return redirect(url_for("iscriviti_success"))
    if not StatusPercorso.query.filter_by(regione=regione).first().iscrizioni:
        stato = StatusPercorso.query.filter_by(regione=regione).first()
        msg = ""
        if stato.data_apertura == "":
            msg = "Le iscrizioni apriranno nei prossimi giorni!"
        elif stato.data_chiusura == "":
            msg = "Le iscrizioni sono momentaneamente chiuse per problemi tecnici, riapriranno a breve!"
        else:
            msg = "Le iscrizioni sono chiuse!<br>Se vuoi registrare una iscrizione tardiva contattaci tramite mail qua sotto!"
        return render_template("iscriviti_chiuse.html", msg=msg, regione=regione)
    return render_template("iscriviti.html", gruppi=gruppi[regione], specialita=specialita, regione=regione)

@app.route("/iscriviti_success")
def iscriviti_success():
    return render_template("iscriviti_success.html")

@app.route("/rel_puglia_file")
@login_required
def rel_puglia_file():
    iscritti = []
    tmp_iscritti=IscrizioniEG.query.filter_by(regione=current_user.regione).filter_by(stato="abilitato")
    for i in tmp_iscritti:
        try:
            stato_rel = RelazioniPuglia.query.filter_by(iscrizioni_id=int(i.id)).first().stato
            if stato_rel == False:
                iscritti.append(i)
        except:
            iscritti.append(i)
    wb = Workbook()
    ws = wb.active
    ws.title = "sq_abilitate"
    #titoli delle colonne
    ws.cell(row=1, column=1).value = "Nome Sq"
    ws.cell(row=1, column=2).value = "Gruppo"
    ws.cell(row=1, column=3).value = "Zona"
    ws.cell(row=1, column=4).value = "Specialità"
    ws.cell(row=1, column=5).value = "Tipo"
    ws.cell(row=1, column=6).value = "Mail Capo Sq"
    ws.cell(row=1, column=7).value = "Mail Capo Rep 1"
    ws.cell(row=1, column=8).value = "Mail Capo Rep 2"
    ws.cell(row=1, column=9).value = "Link"

    for i, iscritto in enumerate(iscritti):
        tmp_riga = i+2
        ws.cell(row=tmp_riga, column=1).value = iscritto.nome
        ws.cell(row=tmp_riga, column=2).value = iscritto.gruppo
        ws.cell(row=tmp_riga, column=3).value = iscritto.zona
        ws.cell(row=tmp_riga, column=4).value = iscritto.specialita
        ws.cell(row=tmp_riga, column=5).value = iscritto.tipo
        ws.cell(row=tmp_riga, column=6).value = iscritto.mail
        ws.cell(row=tmp_riga, column=7).value = iscritto.mail_capo1
        ws.cell(row=tmp_riga, column=8).value = iscritto.mail_capo2
        ws.cell(row=tmp_riga, column=9).value = f"https://guidonciniverdi.pythonanywhere.com/relazione_puglia/{iscritto.id}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="riepilogo.xlsx")

@app.route("/rel_puglia_rep")
@login_required
def rel_puglia_rep():
    iscritti = []
    if current_user.livello == "iabz":
        tmp_iscritti=IscrizioniEG.query.filter_by(regione=current_user.regione).filter_by(stato="abilitato").filter_by(zona=current_user.zona)
    else:
        tmp_iscritti=IscrizioniEG.query.filter_by(regione=current_user.regione).filter_by(stato="abilitato")
    for i in tmp_iscritti:
        try:
            rel = RelazioniPuglia.query.filter_by(iscrizioni_id=int(i.id)).first()
            if rel.stato == True:
                tmp_iscritto = {"nome": i.nome, "gruppo": i.gruppo, "zona": i.zona, "specialita": i.specialita, "tipo": i.tipo, "risposte": rel.dati}
                iscritti.append(tmp_iscritto)
        except:
            pass
    wb = Workbook()
    ws = wb.active
    ws.title = "risposte"
    #titoli delle colonne
    ws.cell(row=1, column=1).value = "Nome Sq"
    ws.cell(row=1, column=2).value = "Gruppo"
    ws.cell(row=1, column=3).value = "Zona"
    ws.cell(row=1, column=4).value = "Specialità"
    ws.cell(row=1, column=5).value = "Tipo"
    for i in range(10):
        ws.cell(row=1, column=6+i).value = f"Risposta {i+1}"

    for i, iscritto in enumerate(iscritti):
        tmp_riga = i+2
        ws.cell(row=tmp_riga, column=1).value = iscritto["nome"]
        ws.cell(row=tmp_riga, column=2).value = iscritto["gruppo"]
        ws.cell(row=tmp_riga, column=3).value = iscritto["zona"]
        ws.cell(row=tmp_riga, column=4).value = iscritto["risposte"]["specialita"]
        ws.cell(row=tmp_riga, column=5).value = iscritto["risposte"]["tipo"]
        for j in range(10):
            ws.cell(row=tmp_riga, column=6+j).value = iscritto["risposte"][f"quest{j+1}"]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="riepilogo.xlsx")

@app.route("/relazione_puglia/<id_sq>", methods=["GET", "POST"])
def relazione_puglia(id_sq):
    sq = IscrizioniEG.query.filter_by(id=int(id_sq)).first()
    try:
        tryrel = RelazioniPuglia.query.filter_by(iscrizioni_id=int(id_sq)).first()
        if tryrel.iscrizioni_id == sq.id and tryrel.stato:
            return render_template("relazione_error.html", nome_sq=sq.nome, gruppo=sq.gruppo)
    except:
        pass
    if request.method == "POST":
        try:
            tryrel = RelazioniPuglia.query.filter_by(iscrizioni_id=int(id_sq)).first()
            if tryrel.iscrizioni_id == sq.id:
                db.session.delete(tryrel)
                db.session.commit()
        except Exception as e:
            print("Errore durante l'eliminazione della relazione:", e)
        try:
            tmp_dati = {
                "specialita": request.form["specialita"],
                "tipo": request.form["conquista_conferma"],
                "quest1": request.form["quest1"],
                "quest2": request.form["quest2"],
                "quest3": request.form["quest3"],
                "quest4": request.form["quest4"],
                "quest5": request.form["quest5"],
                "quest6": request.form["quest6"],
                "quest7": request.form["quest7"],
                "quest8": request.form["quest8"],
                "quest9": request.form["quest9"],
                "quest10": request.form["quest10"]
                }

            relazione = RelazioniPuglia(data=str(datetime.now()), stato=True, iscrizioni_id=int(id_sq), dati=tmp_dati)
            db.session.add(relazione)
            db.session.commit()
        except:
            flash("Invio relazione fallito. Riprovaci!", "warning")
            return redirect(url_for("relazione_puglia", id_sq=id_sq))

        return redirect(url_for("relazione_success"))
    return render_template("relazione_puglia.html", specialita=specialita, nome_sq=sq.nome, gruppo=sq.gruppo)

@app.route("/relazione_success")
def relazione_success():
    return render_template("relazione_success.html")

@app.route("/relazione_delete/<id_sq>")
@login_required
def relazione_delete(id_sq):
    relazione = RelazioniPuglia.query.filter_by(iscrizioni_id=int(id_sq)).first()
    relazione.stato = False
    db.session.commit()
    return redirect(url_for("dettagli", id_iscrizione=id_sq))

@app.route("/notifica", methods=["POST"])
def notifica():
    with open("credenziali_notifiche.json", "r") as f:
        dati = f.read()
    if request.form["api_key"] != json.loads(dati)["api_key"]:
        return {"status": False}
    try:
        request.form["tipologia"]
    except:
        return {"status": False}
    non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").count()
    abilitate = IscrizioniEG.query.filter_by(stato="abilitato").count()
    testo_telegram = f"Da abilitare: {non_abilitate}\nAbilitate: {abilitate}"
    """
    tmp_utenti = User.query.filter_by(livello="admin").all()
    for i in tmp_utenti:
        if non_abilitate > 0:
            manda_telegram(i.telegram_id, "Report REGIONE", testo_telegram)
    if request.form["tipologia"] == "admin":
        return {"status": True}
    """

    tmp_utenti = User.query.filter_by(livello="pattuglia").all()
    for i in tmp_utenti:
        non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").filter_by(regione=request.form["regione"]).count()
        abilitate = IscrizioniEG.query.filter_by(stato="abilitato").filter_by(regione=request.form["regione"]).count()
        testo_telegram = f"Da abilitare: {non_abilitate}\nAbilitate: {abilitate}"
        if non_abilitate > 0:
            manda_telegram(i.telegram_id, f"Report {request.form['regione'].capitalize()}", testo_telegram)

    tmp_utenti = User.query.filter_by(livello="iabr").all()
    for i in tmp_utenti:
        non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").filter_by(regione=request.form["regione"]).count()
        abilitate = IscrizioniEG.query.filter_by(stato="abilitato").filter_by(regione=request.form["regione"]).count()
        testo_telegram = f"Da abilitare: {non_abilitate}\nAbilitate: {abilitate}"
        if non_abilitate > 0:
            manda_telegram(i.telegram_id, f"Report {request.form['regione'].capitalize()}", testo_telegram)
    if request.form["tipologia"] == "regione":
        return {"status": True}

    tmp_utenti = User.query.filter_by(livello="iabz").all()
    for i in tmp_utenti:
        non_abilitate = IscrizioniEG.query.filter_by(stato="da_abilitare").filter_by(zona=i.zona).count()
        abilitate = IscrizioniEG.query.filter_by(stato="abilitato").filter_by(zona=i.zona).count()
        testo_telegram = f"Da abilitare: {non_abilitate}\nAbilitate: {abilitate}"
        if non_abilitate > 0:
            manda_telegram(i.telegram_id, f"Report {i.zona}", testo_telegram)
    return {"status": True}

@app.errorhandler(404)
def page_not_found(e):
    return render_template("errore_generico.html"), 404

@app.errorhandler(405)
def internal_error(e):
    return render_template("errore_generico.html"), 405

@app.errorhandler(500)
def internal_error(e):
    return render_template("errore_generico.html"), 500

if __name__ == "__main__":
    app.run(port=8000, host="0.0.0.0")
